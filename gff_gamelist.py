import tkinter as tk
from tkinter import messagebox
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from datetime import datetime
import time


def scrape_data():
    # Veri çekmek istediğiniz web sitesinin URL'sini belirtin
    url = "https://www.gamesforfarm.com"
    

    # Web sitesine GET isteği gönderin ve yanıtı alın
    response = requests.get(url)

    # İsteğin başarılı olup olmadığını kontrol edin
    if response.status_code == 200:
        # Yanıttaki HTML içeriğini alın
        html_content = response.content

        # BeautifulSoup kullanarak HTML içeriğini analiz edin
        soup = BeautifulSoup(html_content, "html.parser")

        # Ürünleri hedefleyen ana HTML elementini seçin
        product_boxes = soup.find_all("div", class_="product__box")

        # Elde edilen verileri depolamak için bir liste oluşturun
        products = []

        # Her bir ürünü döngüyle gezin
        for product_box in product_boxes:
            # Ürün başlığını çekin
            title_element = product_box.find("div", class_="product__box-title")
            title = title_element.text.strip()

            # Ürün Price(r)ını çekin
            price_element = product_box.find("div", class_="product__box-price")

            # Price(r)ı gereksiz karakterlerden temizleyin
            price_text = price_element.text.replace(",", ".").replace(" р", "")
            price = "".join(filter(lambda x: x.isdigit() or x == ".", price_text))

            # Ürün linkini çekin
            link = title_element.find("a")["href"]

            # Elde edilen verileri bir sözlük olarak depolayın
            product = {
                "Name": title,
                "Price(r)": price,
                "Link": url + link
            }

            # Ürünü ürünler listesine ekleyin
            products.append(product)

        # Verileri pandas DataFrame'e dönüştürün
        df = pd.DataFrame(products)

        # Price(r) sütununu sayısal değerlere dönüştürün
        df["Price(r)"] = df["Price(r)"].astype(float)

        # Price(r)ları küçükten büyüğe doğru sıralayın
        df = df.sort_values("Price(r)")

        # Yinelemeleri kaldırın
        df = df.drop_duplicates()

        # Excel dosyasını oluşturun
        wb = Workbook()
        ws = wb.active

        # Name hücrelerini yazdırın
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title

        # Veri hücrelerini yazdırın
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Kenarlık stili oluşturun
        border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))

        # Tüm hücrelere kenarlık uygulayın
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        # Excel dosyasını kaydedin
        now = datetime.now()
        timestamp = now.strftime("%d%m%Y%H%M")
        file_name = f"data_{timestamp}.xlsx"
        wb.save(file_name)

        print(f"New data {file_name} creating succesfully.")

        # Uygulama tamamlandığında mesaj kutusu göster
        messagebox.showinfo("Scrapping Completed", "Scrapping Completed Succesfully.")

        # Uygulamayı kapat
        time.sleep(3)
        window.destroy()

    else:
        messagebox.showerror("Error", "Can't Acess Website.")


# Tkinter arayüzünü oluştur
window = tk.Tk()
window.title("GFF Scraping")
window.geometry("400x150")
window.resizable(False, False)

# Center the window on the screen
window_width = 400
window_height = 150
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x = int((screen_width / 2) - (window_width / 2))
y = int((screen_height / 2) - (window_height / 2))
window.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Scrapping işlemi başlamadan önceki etiket
start_label = tk.Label(window, text="Scrapping Ready", fg="red", font=("bold", 12))
start_label.pack(pady=20)


# Scraping işlemini başlatan düğme
scrape_button = tk.Button(window, text="Start", command=scrape_data)
scrape_button.pack()

# Scrapping işlemi tamamlandığında gösterilecek etiket
finish_label = tk.Label(window, text="Finish", fg="green", font=("bold", 12))

window.mainloop()
